from collections import defaultdict
from collections.abc import Mapping
import csv
from datetime import datetime
from pathlib import Path
import re
from typing import Final, List, Optional

from loguru import logger
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from inventory.config import env, excel_enabled
from inventory.error.exceptions import NotFoundError
from inventory.model.entity.inventory import Inventory, InventoryType, map_inventory_to_inventory_type
from inventory.model.entity.reserved_item import ReserveInventoryItemType, Reserved_item
from inventory.repository.inventory_repository import InventoryRepository
from inventory.repository.session import get_session
from inventory.repository.slice import Slice
from inventory.repository.pageable import Pageable
from inventory.config.feature_flags import excel_export_enabled  # z. B. True/False-Flag

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, PatternFill, Side

from inventory.client.product.product_service import get_product_by_id


class InventoryReadService:

    def __init__(
        self,
        repository: InventoryRepository,
        session: AsyncSession
        ):
        self._repository = repository
        self._session = session

    async def find_by_id(self, inventory_id: str) -> Inventory:
        logger.debug("inventory_id={}", inventory_id)

        inventory = await self._repository.find_by_id(inventory_id)
        if not inventory:
            raise NotFoundError(inventory_id)
        return inventory

    async def find(
        self,
        filter,
        pageable,
    ):
        logger.debug("find: filter_dict=%s", filter)
        inventorys = await self._repository.find(
            filter_dict=filter,
            pageable=pageable,
        )
        if not inventorys:
            logger.warning("Keine Produkte gefunden mit Filter: %s", filter)
            raise NotFoundError("Keine Produkte mit diesen Filterkriterien gefunden.")
        if excel_export_enabled:
            self._create_export_file(inventorys)
        mapped = [map_inventory_to_inventory_type(p) for p in inventorys]
        return Slice(
            content=mapped,
            total=len(mapped),
            page=pageable.skip,
            size=pageable.limit,
        )

    def _map_to_type(self, item: Reserved_item) -> ReserveInventoryItemType:
        return ReserveInventoryItemType(
            id=item.id,
            customer_id=item.customer_id,
            quantity=item.quantity,
            inventory_id=item.inventory_id,
            created=item.created,
            updated=item.updated,
        )

    def _create_export_file(self,inventorys: List[Inventory]) -> None:
        """Erstellt CSV oder Excel mit Logo und Diagrammen."""

        # 🔢 Startposition
        start_row = 10
        start_col = 2  # B = 2

        # 🔧 Konfiguration
        as_csv: Final = env.EXPORT_FORMAT.lower() == "csv"
        logger.debug("_create_export_file: as csv={}", as_csv)

        # ⏱ Zeitstempel für Dateinamen
        timestamp = datetime.now().strftime("%Y-%m-%d")
        path = Path("exports")
        path.mkdir(parents=True, exist_ok=True)
        file_name = f"{timestamp}.{'csv' if as_csv else 'xlsx'}"
        export_path = path / file_name

        # 🔄 Gemeinsame Zeilenstruktur
        header = [
            "Inventory-ID",
            "SKU-Code",
            "Menge",
            "Einzelpreis",
            "Lagerstatus",
            "Produkt ID",
            "Erstellt",
            "Geändert",
        ]
        rows = [
            [
                str(p.id),
                p.sku_code,
                p.quantity,
                float(p.unit_price),
                p.status.value,
                p.product_id,
                p.created.isoformat(),
                p.updated.isoformat(),
            ]
            for p in inventorys
            if p.unit_price > 0
        ]

        # ➤ CSV
        if as_csv:
            with open(export_path, mode="w", newline="", encoding="utf-8") as csv_file:
                writer = csv.writer(csv_file, delimiter=";")
                writer.writerow(header)
                writer.writerows(rows)
            logger.success("CSV-Export gespeichert unter: %s", export_path)
            return

        # 📊 Excel-Erstellung mit openpyxl
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventar"

        # Rahmen-Stile
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")

        def table_border(row_idx: int, col_idx: int) -> Border:
            """Rahmenregel: außen dick, innen dünn."""

            is_top = row_idx == start_row
            is_bottom = row_idx in {start_row, start_row + len(rows)}
            is_left = col_idx == start_col
            is_right = col_idx == start_col + len(header) - 1

            return Border(
                top=thick if is_top else thin,
                bottom=thick if is_bottom else thin,
                left=thick if is_left else thin,
                right=thick if is_right else thin,
            )

        # Kopfzeile in Zeile 10 (B10)
        for col_offset, title in enumerate(header):
            col_idx = start_col + col_offset
            cell = sheet.cell(row=start_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.border = table_border(start_row, col_idx)

        # 🔴 Zeilen mit Preis > 100 rot markieren
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        for row_offset, row_data in enumerate(rows, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                row_idx = start_row + row_offset
                col_idx = col_index + (start_col - 1)
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = table_border(row_idx, col_idx)

                if col_index == 5:
                    try:
                        if float(cell_value) > 100:
                            cell.fill = red_fill
                    except (ValueError, TypeError):
                        logger.warning("⚠️ Ungültiger Preiswert: {}", cell_value)

        # 🖼️ Branding: Logo in Zelle A1
        # ➤ Logo einfügen (falls vorhanden)
        try:
            logo_path = Path(__file__).parent.parent / "static/logo.png"
            if logo_path.exists():
                logo = ExcelImage(str(logo_path))
                logo.width = 300
                logo.height = 80

                # 🔲 Zellen A1 bis C4 zusammenführen
                sheet.merge_cells("A1:C4")  # Logo-Bereich zusammenführen

                # 📏 Spaltenbreite und Zeilenhöhe für gute Darstellung
                sheet.column_dimensions["A"].width = 20
                sheet.column_dimensions["B"].width = 20
                sheet.column_dimensions["C"].width = 20
                sheet.row_dimensions[1].height = 40
                sheet.row_dimensions[2].height = 40
                sheet.row_dimensions[3].height = 40

                # ➕ Bild in A1 einfügen
                sheet.add_image(logo, "A1")
                print("🔍 Logo hinzugefügt von A1 bis C4")
            else:
                logger.warning("⚠️ Kein Logo gefunden: {}", logo_path)
        except Exception as e:
            logger.warning("⚠️ Fehler beim Logo: {}", str(e))

        # Anzahl Produkte pro Kategorie
        status_count = defaultdict(int)
        status_sum = defaultdict(float)
        status_prices = defaultdict(list)

        for p in inventorys:
            key = p.status.value
            status_count[key] += 1
            status_sum[key] += float(p.unit_price)
            status_prices[key].append((p.product_id, float(p.unit_price)))

        # Kreisdiagramm: Anzahl Produkte pro Kategorie
        sheet_count = workbook.create_sheet("Anzahl je Kategorie")
        sheet_count.append(["Kategorie", "Anzahl"])
        for cat, count in status_count.items():
            sheet_count.append([cat, count])

        pie = PieChart()
        pie.title = "Produkte pro Kategorie"
        data = Reference(
            sheet_count, min_col=2, min_row=1, max_row=len(status_count) + 1
        )
        labels = Reference(
            sheet_count, min_col=1, min_row=2, max_row=len(status_count) + 1
        )
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        sheet_count.add_chart(pie, "E2")

        # Balkendiagramm: Preise pro Kategorie (Summe)
        sheet_sum = workbook.create_sheet("Preise je Kategorie")
        sheet_sum.append(["Kategorie", "Summe (€)"])
        for cat, total in status_sum.items():
            sheet_sum.append([cat, total])

        bar = BarChart()
        bar.title = "Summe der Preise je Kategorie"
        bar.x_axis.title = "Kategorie"
        bar.y_axis.title = "€"
        data = Reference(sheet_sum, min_col=2, min_row=1, max_row=len(status_sum) + 1)
        cats = Reference(sheet_sum, min_col=1, min_row=2, max_row=len(status_sum) + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        sheet_sum.add_chart(bar, "E2")

        # Einzeldiagramme je Kategorie
        for cat, entries in status_prices.items():
            # ws = workbook.create_sheet(f"Preise: {cat[:20]}")
            ws = workbook.create_sheet(self._sanitize_sheet_name(f"Preise {cat}"))
            ws.append(["Produktname", "Preis (€)"])
            for name, unit_price in entries:
                ws.append([name, unit_price])

            chart = BarChart()
            chart.title = f"Preise in Kategorie: {cat}"
            chart.x_axis.title = "Produkt"
            chart.y_axis.title = "Preis (€)"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(entries) + 1)
            names = Reference(ws, min_col=1, min_row=2, max_row=len(entries) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(names)
            ws.add_chart(chart, "D2")

        # 💾 Speichern
        workbook.save(export_path)
        logger.success("📊 Excel gespeichert: {}", export_path)

    def _sanitize_sheet_name(name: str) -> str:
        # Ungültige Excel-Zeichen entfernen oder ersetzen
        invalid_chars = r"[:\\/*?[\]]"
        name = re.sub(invalid_chars, "", name)
        return name[:31]  # Excel-Sheet-Titel dürfen max. 31 Zeichen haben
